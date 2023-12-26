from flask import Flask,render_template,request,redirect,url_for,jsonify
from flask_sqlalchemy import SQLAlchemy 
# import openpyxl
from openpyxl import load_workbook

app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///crud_example.db'
db = SQLAlchemy(app)

class FileItem(db.Model):
   id=db.Column(db.Integer,primary_key=True)
   name=db.Column(db.String(100),nullable=False)
   age=db.Column(db.Integer)

   def __repr__(self) -> str:
      return f"{self.name} {self.age}"



@app.route('/show')
def products():
    allitems=FileItem.query.all()
    print(allitems)
    return 'this is your required data!'

@app.route('/update/<int:id>', methods=['PUT', 'PATCH'])
def update_item(id):
    try:
        data = request.json  
        item = FileItem.query.get(id)

        if item:
            for key, value in data.items():
                setattr(item, key, value)
            db.session.commit()
            return jsonify({'message': f'Item with id {id} updated successfully'}), 200
        else:
            return jsonify({'error': 'Item not found'}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/create', methods=['POST'])
def create_item():
    try:
        data = request.json  # Assuming JSON data in the request
        new_item = FileItem(name=data['name'], 
                            age=data['age']
                            )
        db.session.add(new_item)
        db.session.commit()
        return jsonify({'message': 'Item created successfully'}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/delete/<int:id>', methods=['DELETE'])
def delete_item(id):
    try:
        item = FileItem.query.get(id)
        if item:
            db.session.delete(item)
            db.session.commit()
            return jsonify({'message': f'Deleted item with id {id}'}), 200
        else:
            return jsonify({'error': 'Item not found'}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 500



@app.route('/parse', methods=['POST'])
def parse_excel():
    try:
        data = request.files['file']
        workbook = load_workbook(data)
        sheet = workbook.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            try:
                name, age = row
                # Assuming you have an ApiTask model defined
                data2 = FileItem(name=name, age=age)
                db.session.add(data2)
                # data2.save()
            except ValueError as e:
                print(f"Error unpacking row: {e}")
        db.session.commit()

        return jsonify({'message': 'ok'}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__=="__main__":
     with app.app_context():
      db.create_all()
      app.run(debug=True)